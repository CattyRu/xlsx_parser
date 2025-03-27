import openpyxl
from re import compile, findall, fullmatch
import pandas as pd

from openpyxl.worksheet.worksheet import Worksheet

YEAR = 2023
pd.set_option('display.max_columns', 500)


class ParserXLSX:
    def __init__(self, year: int, file_name_pattern: str = 'files_to_parse/population_belarus_{}.xlsx'):
        self.year = year
        self.file_name = file_name_pattern.format(year)

        self.list_region = []
        self.dict_district_for_district_center = {
            'Минск': 'Минский район',
            'Брест': 'Брестский район',
            'Барановичи': 'Барановичский район',
            'Пинск': 'Пинский район',
            'Витебск': 'Витебский район',
            'Новополоцк': 'Полоцкий район',
            'Гомель': 'Гомельский район',
            'Гродно': 'Гродненский район',
            'Жодино': 'Смолевичский район',
            'Могилев': 'Могилевский район',
            'Бобруйск': 'Бобруйский район',
        }

        self.ws = None
        self.region = ''

        self.df = pd.DataFrame(
            {
                'region': [],  # область
                'district': [],  # район
                'type_atu': [],  # тип административной единицы
                'name': [],  # наименование
                'population_all': [],  # население административной единицы
                'population_city': [],  # городское население административной единицы
                'population_village': []  # население деревень административной единицы
            }
        )

    def parse(self) -> pd.DataFrame:
        self._open_xlsx_file()  # Worksheet
        ws_range, dict_range = self._get_preliminary_cell_range()
        dict_range_regions = self._get_range_for_regions(ws_range)
        print(f'dict with start rows for regions:\t{dict_range_regions}')
        self._get_population_for_regions(dict_range, dict_range_regions)
        self.df['year'] = self.year
        return self.df

    def _open_xlsx_file(self):
        wb = openpyxl.load_workbook(self.file_name)  # Workbook
        self.ws = wb.active  # Worksheet

    def _get_preliminary_cell_range(self) -> (tuple, dict):
        """cropping the worksheet to "Среднегодовая численность населения Республики Беларусь" for easier operation"""
        range = self.ws.dimensions
        last_column = findall(pattern=compile(':(\\D+)'), string=range)[0]
        for row in self.ws:
            for cell in row:
                if isinstance(cell.value, str):
                    if 'Среднегодовая численность населения Республики Беларусь' == cell.value:
                        last_row = cell.row
                        range_ = f'A1:{last_column}{last_row}'
                        ws_range = self.ws[range_]
                        print(f'preliminary range: {range_}')
                        return ws_range, {'l_c': last_column, 'l_r': last_row}
        raise Exception('the document format has changed significantly')

    def _get_range_for_regions(self, ws_range: tuple) -> dict:
        """collecting data on the beginning of each region's section (start_row)"""
        dict_range_regions = {}
        for row in ws_range:
            for cell in row:
                if isinstance(cell.value, str):
                    if 'в разрезе областей и г.Минска' in cell.value:
                        dict_range_regions['Минск'] = cell.row
                    elif 'Брестской области' in cell.value:
                        dict_range_regions['Брестская область'] = cell.row
                    elif 'Витебской области' in cell.value:
                        dict_range_regions['Витебская область'] = cell.row
                    elif 'Гомельской области' in cell.value:
                        dict_range_regions['Гомельская область'] = cell.row
                    elif 'Минской области' in cell.value:
                        dict_range_regions['Минская область'] = cell.row
                    elif 'Гродненской области' in cell.value:
                        dict_range_regions['Гродненская область'] = cell.row
                    elif 'Могилевской области' in cell.value:
                        dict_range_regions['Могилевская область'] = cell.row
        dict_range_regions = dict(sorted(dict_range_regions.items(), key=lambda x: x[1], reverse=False))
        self.list_region = list(dict_range_regions.keys())
        return dict_range_regions

    def _get_population_for_regions(self, dict_range: dict, dict_range_regions: dict) -> None:
        for i in range(0, len(self.list_region)):
            region = self.list_region[i]
            print(f'working with the region "{region}" in the process...')
            start_row = dict_range_regions[region]
            try:
                end_row = dict_range_regions[self.list_region[i + 1]]
            except IndexError:
                end_row = dict_range['l_r']
            range_str = f'A{start_row}:{dict_range['l_c']}{end_row}'
            print(f'\tsection of Worksheet for the region "{region}": {range_str}')
            self.region = region
            ws_range = self.ws[range_str]
            self._processing_data_region(ws_range)
            print(f'\tworking with the region "{region}" has been completed')

    def _processing_data_region(self, ws_range: Worksheet) -> None:
        dict_column_row, num_columns = self._get_range_data_region(ws_range)
        dict_column_row = self._find_regions_name(dict_column_row, num_columns)
        self._read_and_write_data_region(dict_column_row, num_columns)

    @staticmethod
    def _get_range_data_region(ws_range: Worksheet) -> (dict, int):
        dict_column_row = {}
        num_all = 0
        num_city = 0
        num_village = 0
        for row in ws_range:
            for cell in row:
                if isinstance(cell.value, str):
                    if 'все население' in cell.value.replace('\n', ' ').lower():
                        if f'all{num_all}' in dict_column_row:
                            num_all += 1
                        dict_column_row[f'all{num_all}'] = cell.column_letter
                    elif 'городское' in cell.value:
                        if f'city{num_city}' in dict_column_row:
                            num_city += 1
                        dict_column_row[f'city{num_city}'] = cell.column_letter
                    elif 'сельское' in cell.value:
                        if f'village{num_village}' in dict_column_row:
                            num_village += 1
                        dict_column_row[f'village{num_village}'] = cell.column_letter
                        dict_column_row[f'row{num_village}'] = cell.row + 1
        if (num_all == num_city == num_village) is False:
            raise Exception('Something went wrong: the number of columns with the population does not match')
        return dict_column_row, num_all

    def _find_regions_name(self, dict_column_row: dict, num_columns: int) -> dict:
        for i in range(0, num_columns + 1):
            row = dict_column_row[f'row{i}']
            last_column = dict_column_row[f'all{i}']
            row = self.ws[f'A{row}:{last_column}{row}']
            for cell in row[0]:
                if isinstance(cell.value, str):
                    if findall(pattern=compile('г\\.|район'), string=cell.value):
                        dict_column_row[f'region{i}'] = cell.column_letter
        return dict_column_row

    def _read_and_write_data_region(self, dict_column_row: dict, num_columns: int) -> None:
        for i in range(0, num_columns + 1):
            row = dict_column_row[f'row{i}']
            r_names = self.ws[f'{dict_column_row[f'region{i}']}{row}'].value.split('\n')
            all_people = self.ws[f'{dict_column_row[f'all{i}']}{row}'].value.split('\n')
            city_people = self.ws[f"{dict_column_row[f'city{i}']}{row}"].value.split('\n')
            village_people = self.ws[f"{dict_column_row[f'village{i}']}{row}"].value.split('\n')
            if self.region == 'Минск':
                all_people.insert(1, None)
                city_people.insert(1, None)
                village_people.insert(1, None)
            if (len(r_names) == len(all_people) == len(city_people) == len(village_people)) is False:
                raise Exception(f'''Something went wrong: the number of localities and population values do not match
                    number of localities: {len(r_names)},
                    population in location: {len(all_people)},
                    urban population in location: {len(city_people)},
                    rural population in location: {len(village_people)}
                ''')
            if self.region == 'Минск':
                new_df = self._read_data_misk_and_country(r_names, all_people, city_people, village_people)
            else:
                new_df = self._read_data_in_row(r_names, all_people, city_people, village_people)
            self.df = pd.concat([self.df, new_df], ignore_index=True)

    def _read_data_misk_and_country(self, r_names: list, all_people: list, city_people: list, village_people: list
                                    ) -> pd.DataFrame:
        dict_data = {
            'region': [],
            'district': [],
            'type_atu': [],
            'name': [],
            'population_all': [],
            'population_city': [],
            'population_village': []
        }
        for i in range(0, len(r_names)):
            if r_names[i].lower() == 'республика беларусь':
                dict_data['type_atu'].append('страна')
                dict_data['region'].append('-')
                try:
                    dict_data['population_all'].append(self._find_number(all_people[i]))
                    dict_data['population_city'].append(self._find_number(city_people[i]))
                    dict_data['population_village'].append(self._find_number(village_people[i]))
                except TypeError as e:
                    if 'NoneType' in str(e):
                        raise Exception('the document format has changed significantly')
                    else:
                        raise TypeError(f'Something went wrong: unexpected error.\n\t{e}')
                dict_data['district'].append('-')
                dict_data['name'].append('Республика Беларусь')
            elif r_names[i].lower() == 'г.минск':
                dict_data['type_atu'].append('г.')
                dict_data['region'].append('Минская область')
                try:
                    dict_data['population_all'].append(self._find_number(all_people[i]))
                    dict_data['population_city'].append(self._find_number(city_people[i]))
                    dict_data['population_village'].append(self._find_number(village_people[i]))
                except TypeError as e:
                    if 'NoneType' in str(e):
                        raise Exception('the document format has changed significantly')
                    else:
                        raise AttributeError(f'Something went wrong: unexpected error.\n\t{e}')
                dict_data['district'].append(self.dict_district_for_district_center[(r_names[i].split('.'))[-1]])
                dict_data['name'].append('Минск')
        return pd.DataFrame(dict_data)

    def _read_data_in_row(self, r_names: list, all_people: list, city_people: list, village_people: list
                          ) -> pd.DataFrame:
        dict_data = {
            'region': [],
            'district': [],
            'type_atu': [],
            'name': [],
            'population_all': [],
            'population_city': [],
            'population_village': []
        }
        district = ''
        for i in range(0, len(r_names)):
            if r_names[i].lower() == 'всего по области':
                pass
                dict_data['type_atu'].append('область')
                dict_data['region'].append('-')
                dict_data['district'].append('-')
                dict_data['name'].append(self.region)
            elif 'район' in r_names[i]:
                district = r_names[i]
                dict_data['type_atu'].append('район')
                dict_data['region'].append(self.region)
                dict_data['district'].append('-')
                dict_data['name'].append(r_names[i])
            else:
                var = r_names[i].replace(' ', '').split('.')
                type_atu = '.'.join(var[0:-1]) + '.' if len(var) > 2 else var[0] + '.'
                dict_data['type_atu'].append(type_atu)
                dict_data['region'].append(self.region)
                if district:
                    dict_data['district'].append(district)
                else:
                    dict_data['district'].append(self.dict_district_for_district_center[var[-1]])
                dict_data['name'].append(var[-1])
            dict_data['population_all'].append(self._find_number(all_people[i]))
            dict_data['population_city'].append(self._find_number(city_people[i]))
            dict_data['population_village'].append(self._find_number(village_people[i]))
        return pd.DataFrame(dict_data)

    @staticmethod
    def _find_number(str_: str) -> str:
        var = findall(pattern=compile('\\d+'), string=str_)
        if var:
            return ''.join(var)
        else:
            return '-'


if __name__ == '__main__':
    print(ParserXLSX(YEAR).parse())
